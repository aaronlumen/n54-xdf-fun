#!/usr/bin/env python3
# THIS SCRIPT WRITTEN BY AARON SURINA AARON@SURINA.SHOP FOR BMW N54 335i PERFORMANCE TUNING.
# COPYRIGHT 2025 - ALL WRONGS RESERVED.
# FEEL FREE TO DISTRIBUTE ANYTHING THAT YOU'RE ALLOWED TO AND SOME THINGS YOU  MAY NOT BE ALLOWED TO AT FIRST.
# COPYWRONG 2025 SURINA ENTERPRISES - COEUR D'ALENE, IDAHO 
# #
"""
Automated N5X Tuning XDF Update Script
--------------------------------------
Uses an N5X Tuning Spreadsheet (XLSX) to drive high-performance or "race" 
updates in a BMW N54/N55 XDF definition file (I8A0S Custom.xdf, for example).

Key steps:
1. Load XLSX data (parameter name, recommended value, etc.).
2. Parse the XDF (XML).
3. Match spreadsheet rows to relevant XDF elements (e.g., by address or title).
4. Update or rename categories, table titles, max/min bounds, or other attributes.
5. Write a new XDF file with these changes.

You'll need:
  - `openpyxl` Python library (to handle XLSX).
  - `xml.etree.ElementTree` (built-in) for XML manipulation.

Customize the logic to align with your actual spreadsheet columns and the 
specific XDF addresses or item names you want to change.

Adapting to Your Actual Spreadsheet Data

    Identify Column Layout
        In the script, we assume columns [A, B, C, D] might be ['Parameter Name', 'XDF Address', 'NewMax', 'NewTitle'].
        Change for row in sheet.iter_rows(...) logic to match your actual columns. For instance, if your spreadsheet has “Address” in column D, you’d do xdf_address = row[3].

    Matching Strategy
        This example uses mmedaddress from the <EMBEDDEDDATA> of the Z-axis to find a table. You could instead match by <title> text or by <CATEGORYMEM> references—whatever best aligns with how you track them in your spreadsheet.

    Numeric Conversions
        We show how to parse floats for “max” or “min” changes. If your spreadsheet has multiple numeric fields (like new “min,” “max,” or even a new scale factor), incorporate them similarly.

    Updating Other Parts of the XDF
        If you want to rename categories, check <CATEGORY> elements under <XDFHEADER>.
        If you want to modify multiple axes, e.g., the “x” axis or “y” axis, replicate the steps for <XDFAXIS id="x"> or <XDFAXIS id="y">.

Possible Enhancements

    Error Handling: For any row that references an unknown address or invalid numeric data, you can gracefully skip or log the error.
    Conditional Edits: Maybe only raise max if the spreadsheet says “Raise to 300 if supporting mods are installed.”
    Batch vs. Interactive: Currently, the script runs in batch mode. You could add prompts after reading the spreadsheet to confirm changes.
    Flashing BIN: Remember, an XDF is not the actual tune (binary). If your end goal is to directly modify the BIN with these new performance settings, you’d need a separate step to write raw bytes into the BIN, or rely on TunerPro to interpret the updated XDF."""

import xml.etree.ElementTree as ET
import openpyxl
import os
import sys

def load_spreadsheet_data(xlsx_path):
    """
    Reads rows from 'N5X Tuning Spreadsheet.xlsx' and returns them
    in a structured list/dict. Adjust this function based on your actual
    spreadsheet layout.

    Example assumption:
      - 'Parameter Name' in col A
      - 'XDF Address' in col B (e.g., 0x4076A)
      - 'NewMax' in col C (numeric)
      - 'NewTitle' in col D (string, optional)
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    # Suppose we read from the first sheet or a sheet named 'HighPerf'
    sheet_name = wb.sheetnames[0]  # or "HighPerf"
    sheet = wb[sheet_name]

    data = []
    # Assuming row 1 is headers, we start from row 2:
    for row in sheet.iter_rows(min_row=2, values_only=True):
        param_name = row[0]   # e.g. 'Boost Target Table'
        xdf_address = row[1]  # e.g. '0x4076A'
        new_max = row[2]      # e.g. 300.0
        new_title = row[3]    # e.g. 'Race: Max Boost'
        # Or expand for more columns as needed

        if not param_name:  # skip empty lines
            continue

        entry = {
            'param_name': param_name,
            'xdf_address': xdf_address,
            'new_max': new_max,
            'new_title': new_title
        }
        data.append(entry)

    return data

def update_xdf_with_spreadsheet(xdf_path, output_path, spreadsheet_data):
    """
    1. Parse the XDF XML at xdf_path
    2. For each row in spreadsheet_data, attempt to update the relevant 
       table in the XDF (e.g., by matching the 'xdf_address' or table 'title').
    3. Modify attributes like <title>, <max>, <MATH>, etc. for a race tune.
    4. Write the new XDF to output_path
    """

    tree = ET.parse(xdf_path)
    root = tree.getroot()

    # 1) Let's gather all the XDFTABLEs for quick lookups by address or title
    #    We'll do a dictionary: { address -> <XDFTABLE element> } if a 'z-axis' address is found
    #    You might also do a dictionary by the existing <title> text if that's how you prefer to match.
    tables_by_addr = {}

    tables = root.findall('XDFTABLE')
    for table in tables:
        z_axis = table.find("XDFAXIS[@id='z']")
        if z_axis is not None:
            embed = z_axis.find('EMBEDDEDDATA')
            if embed is not None:
                # The 'mmedaddress' attribute often holds the hex offset
                address = embed.get('mmedaddress')
                if address:
                    # Convert both to uppercase for consistency
                    address = address.upper()
                    tables_by_addr[address] = table

    # 2) Process each spreadsheet row
    for row in spreadsheet_data:
        param_name = row['param_name']
        desired_addr = row['xdf_address']
        new_max = row['new_max']
        new_title = row['new_title']

        if not desired_addr:
            continue

        desired_addr = desired_addr.strip().upper()  # '0x4076A' -> '0X4076A'

        # Attempt to locate the table by the desired_addr
        if desired_addr in tables_by_addr:
            table_elem = tables_by_addr[desired_addr]

            # Example A: rename the <title> element
            if new_title:
                title_elem = table_elem.find('title')
                if title_elem is not None:
                    old_title = title_elem.text
                    title_elem.text = new_title
                    print(f"Renamed table title from '{old_title}' to '{new_title}' for address {desired_addr}")

            # Example B: raise <max> in the z-axis if we have a new_max
            if new_max is not None:
                z_axis = table_elem.find("XDFAXIS[@id='z']")
                if z_axis is not None:
                    max_elem = z_axis.find('max')
                    if max_elem is not None:
                        old_max_str = max_elem.text
                        # Convert old max to float for display (if any)
                        try:
                            old_max = float(old_max_str)
                        except:
                            old_max = None
                        max_elem.text = f"{new_max:.2f}"
                        print(f"Updated <max> from {old_max} to {new_max} at address {desired_addr}")

            # You could add more logic: updating <MATH> equations, <min>, or even <EMBEDDEDDATA> attributes.

        else:
            print(f"Warning: No table found at address '{desired_addr}' for param '{param_name}'")

    # 3) Write out the modified file
    tree.write(output_path, encoding="UTF-8", xml_declaration=True)
    print(f"\nSuccessfully wrote updated XDF to: {output_path}")

def main():
    """
    Main flow:
    1. Make sure we have the XDF and the XLSX.
    2. Load the spreadsheet data (the recommended race changes).
    3. Parse + update the XDF accordingly.
    4. Save new XDF.
    """
    xdf_file = "I8A0S Custom.xdf"  # Adjust if needed
    spreadsheet_file = "N5X Tuning Spreadsheet.xlsx"  # Adjust if needed
    output_file = "I8A0S_Custom_RaceMode.xdf"

    if not os.path.isfile(xdf_file):
        print(f"Error: XDF file '{xdf_file}' not found.")
        sys.exit(1)
    if not os.path.isfile(spreadsheet_file):
        print(f"Error: XLSX file '{spreadsheet_file}' not found.")
        sys.exit(1)

    # 1) Load recommended performance data from XLSX
    spreadsheet_data = load_spreadsheet_data(spreadsheet_file)

    # 2) Update XDF
    update_xdf_with_spreadsheet(xdf_file, output_file, spreadsheet_data)

if __name__ == "__main__":
    main()
